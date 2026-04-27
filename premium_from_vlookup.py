import pandas as pd
import numpy as np
from openpyxl import load_workbook

# Read qx and lx from VLOOKUP sheet
df = pd.read_excel("fam_SULT_30_to_49.xlsx", sheet_name="qx_extracted")

# Normalize lx so that lx = 1 at age 30
lx_30 = df.loc[df['age'] == 30, 'lx'].values[0]
df['lx_norm'] = df['lx'] / lx_30 
df['dx'] = df['lx_norm'] * df['qx']

# ADDED: Write lx_norm back to the existing sheet 'qx_extracted' in column D
wb = load_workbook("fam_SULT_30_to_49.xlsx")
ws = wb["qx_extracted"]
ws["D1"] = "lx_norm"
for i, value in enumerate(df['lx_norm'].values, start=2): #Give an index to every element
    ws.cell(row=i, column=4, value=value)
wb.save("fam_SULT_30_to_49.xlsx")
print("Added column D (lx_norm) to sheet 'qx_extracted'")
# END ADDED

# Product assumptions
face = 500000
i = 0.045
v = 1 / (1 + i)
term = 20

lx = df['lx_norm'].values[:term]
dx = df['dx'].values[:term]

t = np.arange(term)
v_t = v ** t
v_d = v ** (t + 1)

A = np.sum(v_d * dx)
a = np.sum(v_t * lx)

net_premium = face * A / a

# Expenses and profit
exp_first = 800
exp_renew = 80
profit_rate = 0.15

numerator = face * A + exp_first + exp_renew * (a - 1)
denominator = a * (1 - profit_rate)
gross_premium = numerator / denominator

print(f"Net premium: {net_premium:.2f}")
print(f"Gross premium: {gross_premium:.2f}")
print(f"a_double_dot: {a:.4f}")
print(f"A_term: {A:.6f}")
print(f"lx at age 30 (normalized): {df.loc[0, 'lx_norm']}")